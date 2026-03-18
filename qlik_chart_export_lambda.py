import csv
import io
import json
import os
import posixpath
import time
from typing import Any, Dict, Optional
from urllib.parse import urlparse

import requests

try:
    import paramiko
except Exception:  # pragma: no cover
    paramiko = None

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


class ConfigError(Exception):
    pass


class QlikExportError(Exception):
    pass


DEFAULT_TIMEOUT = int(os.getenv("HTTP_TIMEOUT_SECONDS", "30"))
DEFAULT_POLL_INTERVAL = float(os.getenv("POLL_INTERVAL_SECONDS", "2"))
DEFAULT_POLL_TIMEOUT = int(os.getenv("POLL_TIMEOUT_SECONDS", "600"))
TMP_DIR = "/tmp"


def env(name: str, default: Optional[str] = None, required: bool = False) -> Optional[str]:
    value = os.getenv(name, default)
    if required and not value:
        raise ConfigError(f"Missing required environment variable: {name}")
    return value


def build_headers(api_key: str) -> Dict[str, str]:
    return {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
        "Accept": "application/json",
        "User-Agent": "qlik-chart-export-lambda/1.0",
    }


def normalize_base_url(base_url: str) -> str:
    return base_url.rstrip("/")


def create_export_payload(
    app_id: str,
    chart_id: str,
    output_type: str,
    temporary_bookmark_id: Optional[str] = None,
    export_options: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    if output_type not in {"xlsx", "csv"}:
        raise ConfigError("output_type must be either 'xlsx' or 'csv'")

    # Qlik Reports API for sense-data-1.0 produces XLSX, not CSV.
    # If CSV is requested we still ask Qlik for XLSX and convert locally.
    payload: Dict[str, Any] = {
        "type": "sense-data-1.0",
        "meta": {
            "exportDeadline": os.getenv("EXPORT_DEADLINE", "P0Y0M0DT0H10M0S"),
            "outputTtl": os.getenv("OUTPUT_TTL", "P0Y0M0DT1H0M0S"),
            "tags": ["lambda", "chart-export"],
        },
        "senseDataTemplate": {
            "appId": app_id,
            "id": chart_id,
            "selectionType": "selectionsByState",
            "selectionsByState": {},
            "exportOptions": export_options or {
                "showTitles": False,
                "showTotals": False,
                "showSelections": False,
            },
        },
        "output": {
            "outputId": f"Chart_{output_type}",
            "type": "xlsx",
            "xlsxOutput": {},
        },
    }

    if temporary_bookmark_id:
        payload["senseDataTemplate"]["selectionType"] = "temporaryBookmarkV2"
        payload["senseDataTemplate"]["temporaryBookmarkV2"] = {"id": temporary_bookmark_id}
        payload["senseDataTemplate"].pop("selectionsByState", None)

    return payload


def request_export(base_url: str, headers: Dict[str, str], payload: Dict[str, Any]) -> Dict[str, Any]:
    url = f"{normalize_base_url(base_url)}/api/v1/reports"
    response = requests.post(url, headers=headers, json=payload, timeout=DEFAULT_TIMEOUT)
    response.raise_for_status()
    body = response.json()
    request_id = body.get("requestId")
    status_url = response.headers.get("Location") or f"{normalize_base_url(base_url)}/api/v1/reports/{request_id}/status"
    outputs_url = body.get("outputsUrl") or f"{normalize_base_url(base_url)}/api/v1/reports/{request_id}/outputs"

    if not request_id:
        raise QlikExportError(f"Qlik did not return requestId. Response: {body}")

    return {
        "request_id": request_id,
        "status_url": status_url,
        "outputs_url": outputs_url,
        "initial_response": body,
    }


def poll_until_done(headers: Dict[str, str], status_url: str, outputs_url: str) -> Dict[str, Any]:
    deadline = time.time() + DEFAULT_POLL_TIMEOUT
    last_status = None

    while time.time() < deadline:
        status_resp = requests.get(status_url, headers=headers, timeout=DEFAULT_TIMEOUT)
        status_resp.raise_for_status()
        status_body = status_resp.json()
        last_status = status_body.get("status")

        if last_status == "done":
            break
        if last_status in {"failed", "aborted"}:
            outputs = get_outputs(headers, outputs_url)
            raise QlikExportError(
                f"Qlik export ended with status '{last_status}'. Outputs: {json.dumps(outputs)}"
            )

        time.sleep(DEFAULT_POLL_INTERVAL)
    else:
        raise TimeoutError(f"Timed out waiting for report generation. Last status: {last_status}")

    outputs = get_outputs(headers, outputs_url)
    for output in outputs.get("data", []):
        if output.get("status") == "done" and output.get("location"):
            return output

    raise QlikExportError(f"No completed output returned from Qlik. Outputs: {json.dumps(outputs)}")


def get_outputs(headers: Dict[str, str], outputs_url: str) -> Dict[str, Any]:
    resp = requests.get(outputs_url, headers=headers, timeout=DEFAULT_TIMEOUT)
    resp.raise_for_status()
    return resp.json()


def download_file(headers: Dict[str, str], download_url: str, target_path: str) -> str:
    with requests.get(download_url, headers=headers, timeout=DEFAULT_TIMEOUT, stream=True) as response:
        response.raise_for_status()
        with open(target_path, "wb") as file_handle:
            for chunk in response.iter_content(chunk_size=1024 * 1024):
                if chunk:
                    file_handle.write(chunk)
    return target_path


def convert_xlsx_to_csv(xlsx_path: str, csv_path: str, sheet_name: Optional[str] = None) -> str:
    if load_workbook is None:
        raise ConfigError("openpyxl is required when output_type=csv")

    workbook = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]

    with open(csv_path, "w", newline="", encoding="utf-8") as csv_file:
        writer = csv.writer(csv_file)
        for row in sheet.iter_rows(values_only=True):
            writer.writerow(["" if value is None else value for value in row])

    return csv_path


def upload_sftp(
    local_path: str,
    host: str,
    username: str,
    password: str,
    remote_path: str,
    port: int = 22,
) -> None:
    if paramiko is None:
        raise ConfigError("paramiko is required for SFTP upload")

    transport = paramiko.Transport((host, port))
    try:
        transport.connect(username=username, password=password)
        sftp = paramiko.SFTPClient.from_transport(transport)
        try:
            ensure_remote_dirs(sftp, posixpath.dirname(remote_path))
            sftp.put(local_path, remote_path)
        finally:
            sftp.close()
    finally:
        transport.close()


def ensure_remote_dirs(sftp: "paramiko.SFTPClient", remote_dir: str) -> None:
    if not remote_dir or remote_dir == "/":
        return

    current = ""
    for part in remote_dir.split("/"):
        if not part:
            continue
        current = f"{current}/{part}"
        try:
            sftp.stat(current)
        except FileNotFoundError:
            sftp.mkdir(current)


def default_filename(chart_id: str, output_type: str) -> str:
    ts = time.strftime("%Y%m%d-%H%M%S", time.gmtime())
    return f"qlik-{chart_id}-{ts}.{output_type}"


def send_slack_notification(text: str) -> None:
    webhook_url = env("SLACK_WEBHOOK_URL")
    if not webhook_url:
        return
    try:
        requests.post(webhook_url, json={"text": text}, timeout=10)
    except Exception:
        pass


def lambda_handler(event: Dict[str, Any], context: Any) -> Dict[str, Any]:
    """
    Example event:
    {
      "qlik_base_url": "https://tenant.eu.qlikcloud.com",
      "app_id": "c3e24281-32ba-48d7-944d-74d537336e29",
      "chart_id": "tPZH",
      "output_type": "xlsx",
      "temporary_bookmark_id": "Temporary\\b4ba78ba-07c3-4b08-bd8d-62fc3112b342",
      "export_options": {
        "showTitles": false,
        "showTotals": false,
        "showSelections": false
      },
      "sftp": {
        "enabled": true,
        "host": "sftp.example.com",
        "port": 22,
        "username": "myuser",
        "password": "mypassword",
        "remote_path": "/incoming/qlik/chart.xlsx"
      }
    }
    """

    try:
        return _run(event)
    except Exception as exc:
        send_slack_notification(f":x: *Qlik chart export FAILED*\nError: {exc}")
        raise


def _run(event: Dict[str, Any]) -> Dict[str, Any]:
    qlik_base_url = event.get("qlik_base_url") or env("QLIK_BASE_URL", required=True)
    qlik_api_key = event.get("qlik_api_key") or env("QLIK_API_KEY", required=True)
    app_id = event.get("app_id") or env("QLIK_APP_ID", required=True)
    chart_id = event.get("chart_id") or env("QLIK_CHART_ID", required=True)
    output_type = (event.get("output_type") or env("OUTPUT_TYPE", "xlsx")).lower()
    temporary_bookmark_id = event.get("temporary_bookmark_id") or env("QLIK_TEMPORARY_BOOKMARK_ID")
    export_options = event.get("export_options")
    csv_sheet_name = event.get("csv_sheet_name") or env("CSV_SHEET_NAME")

    headers = build_headers(qlik_api_key)
    payload = create_export_payload(
        app_id=app_id,
        chart_id=chart_id,
        output_type=output_type,
        temporary_bookmark_id=temporary_bookmark_id,
        export_options=export_options,
    )

    export_request = request_export(qlik_base_url, headers, payload)
    completed_output = poll_until_done(
        headers=headers,
        status_url=export_request["status_url"],
        outputs_url=export_request["outputs_url"],
    )

    filename = event.get("filename") or default_filename(chart_id, output_type)
    xlsx_filename = filename if output_type == "xlsx" else filename.rsplit(".", 1)[0] + ".xlsx"
    xlsx_local_path = os.path.join(TMP_DIR, xlsx_filename)

    download_file(headers, completed_output["location"], xlsx_local_path)

    final_path = xlsx_local_path
    final_filename = xlsx_filename

    if output_type == "csv":
        csv_filename = filename if filename.endswith(".csv") else filename.rsplit(".", 1)[0] + ".csv"
        csv_local_path = os.path.join(TMP_DIR, csv_filename)
        convert_xlsx_to_csv(xlsx_local_path, csv_local_path, sheet_name=csv_sheet_name)
        final_path = csv_local_path
        final_filename = csv_filename

    sftp_config = event.get("sftp", {})
    sftp_enabled = bool(sftp_config.get("enabled") or env("SFTP_ENABLED", "false").lower() == "true")

    sftp_result: Dict[str, Any] = {"enabled": sftp_enabled}
    if sftp_enabled:
        host = sftp_config.get("host") or env("SFTP_HOST", required=True)
        username = sftp_config.get("username") or env("SFTP_USERNAME", required=True)
        password = sftp_config.get("password") or env("SFTP_PASSWORD", required=True)
        port = int(sftp_config.get("port") or env("SFTP_PORT", "22"))
        remote_path = sftp_config.get("remote_path") or env("SFTP_REMOTE_PATH", required=True)

        if remote_path.endswith("/"):
            remote_path = posixpath.join(remote_path, final_filename)

        upload_sftp(
            local_path=final_path,
            host=host,
            username=username,
            password=password,
            remote_path=remote_path,
            port=port,
        )
        sftp_result.update({
            "uploaded": True,
            "host": host,
            "port": port,
            "remote_path": remote_path,
        })

    result = {
        "ok": True,
        "request_id": export_request["request_id"],
        "status": completed_output.get("status"),
        "qlik_download_url": completed_output.get("location"),
        "local_file": final_path,
        "file_name": final_filename,
        "sftp": sftp_result,
    }

    sftp_info = f"uploaded to `{sftp_result.get('remote_path')}`" if sftp_result.get("uploaded") else "SFTP disabled"
    send_slack_notification(
        f":white_check_mark: *Qlik chart export succeeded*\n"
        f"File: `{final_filename}`\n"
        f"SFTP: {sftp_info}"
    )

    return result


if __name__ == "__main__":
    sample_event = {
        "qlik_base_url": env("QLIK_BASE_URL"),
        "qlik_api_key": env("QLIK_API_KEY"),
        "app_id": env("QLIK_APP_ID"),
        "chart_id": env("QLIK_CHART_ID"),
        "output_type": env("OUTPUT_TYPE", "xlsx"),
        "temporary_bookmark_id": env("QLIK_TEMPORARY_BOOKMARK_ID"),
        "sftp": {
            "enabled": env("SFTP_ENABLED", "false").lower() == "true",
            "host": env("SFTP_HOST"),
            "port": int(env("SFTP_PORT", "22")),
            "username": env("SFTP_USERNAME"),
            "password": env("SFTP_PASSWORD"),
            "remote_path": env("SFTP_REMOTE_PATH", "/incoming/"),
        },
    }
    print(json.dumps(lambda_handler(sample_event, None), indent=2))
